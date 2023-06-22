from django.shortcuts import render
from django.conf import settings
import openpyxl
from datetime import datetime
import os
from .models import SectorType, Society, SECTOR_TYPE
import datetime
from collections import Counter

# Create your views here.


def dashboard(request):
    # excel_file_path = os.path.join(str(settings.BASE_DIR), 'dashboard', 'static', 'attachement.xlsx')
    # result = extract_tables(excel_file_path)
    get_year = int(request.GET.get('year', '2023'))

    total_sector_type = SectorType.objects.filter().count()
    unique_states_count = Society.objects.filter(registration_date__year = get_year).values('state').distinct().count()
    total_society_names = Society.objects.filter(registration_date__year = get_year).count()
    today = datetime.date.today()
    current_month_registrations = Society.objects.filter(
        registration_date__month=today.month,
        registration_date__year=today.year
    ).count()
    society_data = Society.objects.filter()
    sector = SectorType.objects.filter()
    state = Society.objects.values('state').distinct()
    current_month_data = Society.objects.filter(
        registration_date__month=today.month,
        registration_date__year=today.year
    )


    xValues = [datetime.date(get_year, i, 1).strftime('%B') for i in range(1, 13)]

    # Get the total number of registrations per month
    registrations_per_month = Counter()

    for society in society_data:
        registration_date = society.registration_date
        if registration_date is not None and registration_date.year == get_year:
            registrations_per_month[registration_date.month] += 1

    # Populate yValues with the total number of registrations for each month
    yValues = [registrations_per_month.get(month, 0) for month in range(1, 13)]


   # Assuming society_data is a list of Society objects with sector_type and registration_date fields
    society_data_sector = Society.objects.filter(registration_date__year=get_year)

    # Get the sector names
    pxValues = [sector.sector_name for sector in SectorType.objects.all()]

    # Get the total number of registrations per sector
    registrations_per_sector = Counter(society.sector_type_id for society in society_data_sector)

    # Populate yValues with the total number of registrations for each sector
    pyValues = [registrations_per_sector.get(sector.id, 0) for sector in SectorType.objects.all()]
    return render(request, 'index.html', locals())



def get_sector_type_value(sector_type):
    for value, name in SECTOR_TYPE:
        try:
            if name.lower() == sector_type.lower():
                return value
        except:
            return None
    return None

def extract_tables(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheets = workbook.sheetnames
    data = []
    header = ['sr_no', 'society_name', 'address', 'state', 'district', 'registration_date', 'area_operation',
              'sector_type']
    for sheet_name in sheets:
        sheet = workbook[sheet_name]
        count = 0
        for row in sheet.iter_rows(values_only=True, min_col=1, max_col=8):  # Extract columns A to I
            if count == 0:
                count = 1
                continue
            main_row = list(row)
            registration_date = main_row[header.index('registration_date')]
            if isinstance(registration_date, datetime.datetime):
                main_row[header.index('registration_date')] = registration_date.strftime('%Y-%m-%d')

            sector_type = main_row[header.index('sector_type')]
            sector_type_value = get_sector_type_value(sector_type)
            if sector_type_value is not None:
                sector_obj = SectorType.objects.get(id=sector_type_value)
                main_row[header.index('sector_type')] = sector_obj
            print(main_row)
            dictionary_obj = dict(zip(header, main_row))
            Society.objects.create(**dictionary_obj)
            data.append(dictionary_obj)
    return data

